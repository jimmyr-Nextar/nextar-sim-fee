import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Nextar SIM Fee Report",
    page_icon="📋",
    layout="wide",
)

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Base */
    [data-testid="stAppViewContainer"] { background: #F7F8FA; }
    [data-testid="stSidebar"] { background: #1F3864; }
    [data-testid="stSidebar"] * { color: #FFFFFF !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stTextInput label { color: #BDD7EE !important; font-size: 0.8rem; }

    /* Header */
    .app-header {
        background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
        padding: 2rem 2.5rem 1.5rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .app-header h1 { color: white; font-size: 1.8rem; font-weight: 700; margin: 0; }
    .app-header p { color: #BDD7EE; margin: 0.3rem 0 0; font-size: 0.95rem; }

    /* Section cards */
    .section-card {
        background: white;
        border-radius: 10px;
        padding: 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }
    .section-title {
        font-size: 1rem;
        font-weight: 700;
        color: #1F3864;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #BDD7EE;
    }

    /* Status badges */
    .badge-ok { background:#E2EFDA; color:#375623; padding:3px 10px; border-radius:20px; font-size:0.8rem; font-weight:600; }
    .badge-warn { background:#FFF3CD; color:#856404; padding:3px 10px; border-radius:20px; font-size:0.8rem; font-weight:600; }
    .badge-err { background:#F8D7DA; color:#721C24; padding:3px 10px; border-radius:20px; font-size:0.8rem; font-weight:600; }

    /* Fee table */
    .fee-table { width:100%; border-collapse:collapse; font-size:0.9rem; }
    .fee-table th { background:#2E75B6; color:white; padding:8px 12px; text-align:left; }
    .fee-table td { padding:7px 12px; border-bottom:1px solid #E2E8F0; }
    .fee-table tr:last-child td { border-bottom:none; }
    .fee-table tr:nth-child(even) td { background:#F7F8FA; }

    /* COO entry */
    .coo-entry {
        background: #DEEAF1;
        border-left: 4px solid #2E75B6;
        border-radius: 0 8px 8px 0;
        padding: 1rem;
        margin-bottom: 0.75rem;
    }

    /* Summary boxes */
    .metric-box {
        background: white;
        border-radius: 10px;
        padding: 1.2rem 1.5rem;
        border: 1px solid #E2E8F0;
        text-align: center;
    }
    .metric-label { font-size: 0.8rem; color: #6B7280; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }
    .metric-value { font-size: 1.8rem; font-weight: 700; color: #1F3864; }
    .metric-sub { font-size: 0.85rem; color: #6B7280; }

    /* Generate button */
    .stButton > button {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 700;
        font-size: 1rem;
        width: 100%;
        transition: opacity 0.2s;
    }
    .stButton > button:hover { opacity: 0.9; }

    /* Hide streamlit chrome */
    #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <h1>📋 Nextar SIM Fee Report</h1>
    <p>BYOD Monthly Report Generator — Sub Dealer & Preferred Only</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR — Report Settings
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Report Settings")
    st.markdown("---")

    report_month = st.selectbox("Report Month", [
        "January","February","March","April","May","June",
        "July","August","September","October","November","December"
    ], index=date.today().month - 2 if date.today().month > 1 else 11)

    report_year = st.number_input("Report Year", min_value=2024, max_value=2030,
                                   value=date.today().year, step=1)

    st.markdown("---")
    st.markdown("### 💰 Fee Schedule")
    st.caption("Edit tiers below. Changes apply to current run only.")

    t1_max = st.number_input("Tier 1 max (< this = $0)", value=4.00, step=0.01, format="%.2f")
    t2_max = st.number_input("Tier 2 max ($5 fee up to)", value=7.00, step=0.01, format="%.2f")
    t3_max = st.number_input("Tier 3 max ($10 fee up to)", value=11.00, step=0.01, format="%.2f")

    st.caption(f"""
    **Current tiers:**
    - < ${t1_max:.2f} → $0
    - ${t1_max:.2f}–${t2_max:.2f} → $5
    - ${t2_max:.2f}–${t3_max:.2f} → $10
    - > ${t3_max:.2f} → $15
    - JKME doors → $5 (always)
    """)

    st.markdown("---")
    st.markdown("### 📦 Return Credits")
    phone_credit = st.number_input("Phone credit ($)", value=20, step=1)
    bts_credit = st.number_input("BTS credit ($)", value=9, step=1)

# ─────────────────────────────────────────────
# HELPER — Fee calc
# ─────────────────────────────────────────────
def calc_fee(comp, is_jkme, t1, t2, t3):
    if is_jkme:
        return 5
    if pd.isna(comp) or comp < t1:
        return 0
    elif comp <= t2:
        return 5
    elif comp <= t3:
        return 10
    else:
        return 15

# ─────────────────────────────────────────────
# SECTION 1 — File Uploads
# ─────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">📁 Upload Files</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    byod_file = st.file_uploader("BYOD MCS Report (current month)", type=["xlsx"], key="byod")
    prev_door_file = st.file_uploader("Door Info — Previous Month", type=["xlsx"], key="prev_door")
with col2:
    returns_file = st.file_uploader("Remorse Returns (current month)", type=["xlsx"], key="returns")
    curr_door_file = st.file_uploader("Door Info — Current Month", type=["xlsx"], key="curr_door")

st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SECTION 2 — COO Exceptions
# ─────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">🔄 Change of Ownership Exceptions</div>', unsafe_allow_html=True)
st.caption("Add entries only when a store changed ownership mid-month. Leave empty if none.")

if "coo_entries" not in st.session_state:
    st.session_state.coo_entries = []

def add_coo():
    st.session_state.coo_entries.append({
        "door_code": "",
        "old_dealer": "",
        "old_type": "Direct",
        "new_dealer": "",
        "new_type": "Preferred",
        "effective_date": date.today(),
    })

def remove_coo(i):
    st.session_state.coo_entries.pop(i)

if st.button("＋ Add COO Exception"):
    add_coo()

for i, entry in enumerate(st.session_state.coo_entries):
    st.markdown(f'<div class="coo-entry">', unsafe_allow_html=True)
    st.markdown(f"**Exception #{i+1}**")
    c1, c2, c3 = st.columns([1, 2, 1])
    with c1:
        st.session_state.coo_entries[i]["door_code"] = st.text_input(
            "Door Code", value=entry["door_code"], key=f"door_{i}")
    with c2:
        st.session_state.coo_entries[i]["effective_date"] = st.date_input(
            "Effective Date (new owner starts)", value=entry["effective_date"], key=f"date_{i}")
    with c3:
        if st.button(f"✕ Remove", key=f"remove_{i}"):
            remove_coo(i)
            st.rerun()

    c4, c5, c6, c7 = st.columns(4)
    with c4:
        st.session_state.coo_entries[i]["old_dealer"] = st.text_input(
            "Old Dealer Name", value=entry["old_dealer"], key=f"old_name_{i}")
    with c5:
        st.session_state.coo_entries[i]["old_type"] = st.selectbox(
            "Old Account Type", ["Direct","Sub","Preferred"],
            index=["Direct","Sub","Preferred"].index(entry["old_type"]), key=f"old_type_{i}")
    with c6:
        st.session_state.coo_entries[i]["new_dealer"] = st.text_input(
            "New Dealer Name", value=entry["new_dealer"], key=f"new_name_{i}")
    with c7:
        st.session_state.coo_entries[i]["new_type"] = st.selectbox(
            "New Account Type", ["Direct","Sub","Preferred"],
            index=["Direct","Sub","Preferred"].index(entry["new_type"]), key=f"new_type_{i}")
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SECTION 3 — Generate
# ─────────────────────────────────────────────
all_uploaded = all([byod_file, returns_file, prev_door_file, curr_door_file])

if not all_uploaded:
    missing = []
    if not byod_file: missing.append("BYOD MCS Report")
    if not returns_file: missing.append("Remorse Returns")
    if not prev_door_file: missing.append("Previous Door Info")
    if not curr_door_file: missing.append("Current Door Info")
    st.info(f"⬆️ Please upload: {', '.join(missing)}")
else:
    st.markdown('<div class="section-card"><div class="section-title">🚀 Generate Report</div>', unsafe_allow_html=True)

    if st.button(f"Generate {report_month} {report_year} SIM Fee Report"):
        with st.spinner("Processing data..."):
            try:
                # ── Load files ──
                byod = pd.read_excel(byod_file, dtype=str)
                ret = pd.read_excel(returns_file, dtype=str)
                prev_doors = pd.read_excel(prev_door_file, dtype=str)
                curr_doors = pd.read_excel(curr_door_file, dtype=str)

                # Normalize door info columns
                for df in [prev_doors, curr_doors]:
                    if 'Nextar Market ID' in df.columns:
                        df.rename(columns={'Nextar Market ID': 'Market'}, inplace=True)
                    if 'Account type' in df.columns:
                        df.rename(columns={'Account type': 'Account Type'}, inplace=True)

                all_doors = pd.concat([curr_doors, prev_doors]).drop_duplicates(subset='Dealer Code', keep='first')

                # ── BYOD processing ──
                byod['Comp'] = pd.to_numeric(byod['Transaction Amount2'], errors='coerce')
                byod['TxDate'] = pd.to_datetime(byod['Transaction Date'], errors='coerce')

                merged = byod.merge(
                    all_doors[['Dealer Code','Market','Account Type','Location Name','RT2 ID','Street Address','City','State','ZIP']],
                    left_on='Door Code', right_on='Dealer Code', how='left'
                )

                # ── Apply COO exceptions ──
                coo_log = []
                for entry in st.session_state.coo_entries:
                    if not entry["door_code"].strip():
                        continue
                    dc = entry["door_code"].strip()
                    eff = pd.Timestamp(entry["effective_date"])
                    mask = merged['Door Code'] == dc

                    before_mask = mask & (merged['TxDate'] < eff)
                    after_mask = mask & (merged['TxDate'] >= eff)

                    # Set account type per period
                    merged.loc[before_mask, 'Account Type'] = entry["old_type"]
                    merged.loc[before_mask, 'Location Name'] = entry["old_dealer"]
                    merged.loc[after_mask, 'Account Type'] = entry["new_type"]
                    merged.loc[after_mask, 'Location Name'] = entry["new_dealer"]

                    before_count = before_mask.sum()
                    after_count = after_mask.sum()
                    coo_log.append({
                        "Door Code": dc,
                        "Old Dealer": entry["old_dealer"],
                        "Old Type": entry["old_type"],
                        "New Dealer": entry["new_dealer"],
                        "New Type": entry["new_type"],
                        "Effective Date": entry["effective_date"].strftime("%m-%d-%Y"),
                        "Txns Before (Old Owner)": before_count,
                        "Txns After (New Owner)": after_count,
                    })

                # ── Filter Sub/Preferred ──
                merged = merged[merged['Account Type'].str.strip().str.lower().isin(['sub','preferred'])].copy()
                merged['Account Type'] = merged['Account Type'].str.strip()
                merged['Is_JKME'] = merged['Location Name'].str.contains('JKME', case=False, na=False)

                # ── Calculate fees ──
                merged['BYOD Fee'] = merged.apply(
                    lambda r: calc_fee(r['Comp'], r['Is_JKME'], t1_max, t2_max, t3_max), axis=1
                )
                for col in ['Comp MRC','Transaction Amount','Transaction Amount2','Commission MRC']:
                    if col in merged.columns:
                        merged[col] = pd.to_numeric(merged[col], errors='coerce')

                # ── Returns ──
                ret_merged = ret.merge(
                    all_doors[['Dealer Code','Market','Account Type','Location Name']],
                    left_on='Door Code', right_on='Dealer Code', how='left'
                )
                ret_merged = ret_merged[ret_merged['Account Type'].str.strip().str.lower().isin(['sub','preferred'])].copy()
                if 'Device Type' in ret_merged.columns:
                    ret_merged['Return Credit'] = ret_merged['Device Type'].apply(
                        lambda x: phone_credit if str(x).strip().lower() == 'phone'
                        else (bts_credit if str(x).strip().lower() == 'bts' else 0)
                    )
                else:
                    ret_merged['Return Credit'] = 0
                    st.warning("⚠️ 'Device Type' column not found in Returns file — return credits set to $0.")

                ret_by_market = ret_merged.groupby('Market')['Return Credit'].sum()
                ret_by_dealer = ret_merged.groupby(['Market','Location Name'])['Return Credit'].sum()
                ret_by_door = ret_merged.groupby('Door Code')['Return Credit'].sum()

                # ── Summaries ──
                market_fees = merged.groupby('Market')['BYOD Fee'].sum()
                market_acts = merged.groupby('Market')['BYOD Fee'].count()
                market_summary = pd.DataFrame({'BYOD_Fee': market_fees, 'Activations': market_acts}).reset_index()
                market_summary['Returns'] = market_summary['Market'].map(ret_by_market).fillna(0).astype(int)
                market_summary['Net_Fee'] = market_summary['BYOD_Fee'] - market_summary['Returns']
                market_summary = market_summary.sort_values('BYOD_Fee', ascending=False)

                dealer_fees = merged.groupby(['Market','Account Type','Location Name','Is_JKME']).agg(
                    Doors=('Door Code','nunique'), Activations=('Comp','count'), Total_SIM_Fee=('BYOD Fee','sum')
                ).reset_index()
                dealer_fees['Returns'] = dealer_fees.apply(
                    lambda r: ret_by_dealer.get((r['Market'], r['Location Name']), 0), axis=1
                )
                dealer_fees['Net_Fee'] = dealer_fees['Total_SIM_Fee'] - dealer_fees['Returns']
                dealer_fees = dealer_fees.sort_values(['Market','Location Name'])

                door_fees = merged.groupby(
                    ['Door Code','Market','Account Type','Location Name','Street Address','City','State','ZIP','RT2 ID','Is_JKME']
                ).agg(Activations=('Comp','count'), BYOD_Fee=('BYOD Fee','sum')).reset_index()
                door_fees['Returns'] = door_fees['Door Code'].map(ret_by_door).fillna(0).astype(int)
                door_fees['Total_Fee'] = door_fees['BYOD_Fee'] - door_fees['Returns']
                door_fees = door_fees.sort_values(['Market','Location Name','Door Code'])

                # ── Preview metrics ──
                total_fee = int(market_summary['BYOD_Fee'].sum())
                total_ret = int(market_summary['Returns'].sum())
                total_net = int(market_summary['Net_Fee'].sum())
                total_acts = int(market_summary['Activations'].sum())
                total_doors = len(door_fees)
                total_dealers = len(dealer_fees)

                st.markdown("---")
                m1, m2, m3, m4, m5, m6 = st.columns(6)
                for col, label, val, sub in [
                    (m1, "Activations", f"{total_acts:,}", "transactions"),
                    (m2, "Doors", f"{total_doors:,}", "unique doors"),
                    (m3, "Dealers", f"{total_dealers:,}", "unique dealers"),
                    (m4, "SIM Fee", f"${total_fee:,}", "before returns"),
                    (m5, "Returns", f"${total_ret:,}", "credits"),
                    (m6, "Net SIM Fee", f"${total_net:,}", "final amount"),
                ]:
                    col.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-label">{label}</div>
                        <div class="metric-value">{val}</div>
                        <div class="metric-sub">{sub}</div>
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown("#### Market Summary Preview")
                preview = market_summary.copy()
                preview.columns = ['Market','Activations','SIM Fee','Returns','Net SIM Fee']
                preview['SIM Fee'] = preview['SIM Fee'].apply(lambda x: f"${x:,.0f}")
                preview['Returns'] = preview['Returns'].apply(lambda x: f"${x:,.0f}")
                preview['Net SIM Fee'] = preview['Net SIM Fee'].apply(lambda x: f"${x:,.0f}")
                st.dataframe(preview, use_container_width=True, hide_index=True)

                if coo_log:
                    st.markdown("#### COO Exceptions Applied")
                    st.dataframe(pd.DataFrame(coo_log), use_container_width=True, hide_index=True)

                # ─────────────────────────────────────────────
                # BUILD EXCEL
                # ─────────────────────────────────────────────
                DARK_BLUE="1F3864"; MED_BLUE="2E75B6"; LIGHT_BLUE="BDD7EE"
                VERY_LIGHT="DEEAF1"; WHITE="FFFFFF"; GOLD="FFD966"; GRAY="F2F2F2"
                LIGHT_GREEN="E2EFDA"

                def thin_border():
                    s = Side(style='thin', color='BFBFBF')
                    return Border(left=s, right=s, top=s, bottom=s)

                def hdr(cell, text, bg=MED_BLUE, fg=WHITE, size=10, wrap=False):
                    cell.value = text
                    cell.font = Font(bold=True, color=fg, size=size, name='Arial')
                    cell.fill = PatternFill('solid', start_color=bg)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)

                def dcell(cell, value, bold=False, bg=None, fg="000000", center=False, num_fmt=None):
                    cell.value = value
                    cell.font = Font(bold=bold, color=fg, name='Arial', size=10)
                    if bg: cell.fill = PatternFill('solid', start_color=bg)
                    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
                    if num_fmt: cell.number_format = num_fmt

                def border_row(ws, row, c1, c2):
                    for c in range(c1, c2+1):
                        ws.cell(row, c).border = thin_border()

                def title_row(ws, text, ncols, row=1):
                    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
                    t = ws.cell(row, 1)
                    t.value = text
                    t.font = Font(bold=True, size=15, color=WHITE, name='Arial')
                    t.fill = PatternFill('solid', start_color=DARK_BLUE)
                    t.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[row].height = 28

                def subtitle_row(ws, text, ncols, row=2):
                    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
                    s = ws.cell(row, 1)
                    s.value = text
                    s.font = Font(bold=True, size=10, color=WHITE, name='Arial')
                    s.fill = PatternFill('solid', start_color=MED_BLUE)
                    s.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[row].height = 18

                wb = Workbook()
                report_label = f"{report_month} {report_year}"

                # ── Sheet 1: Market ──
                ws1 = wb.active; ws1.title = "Market"; ws1.sheet_view.showGridLines = False
                title_row(ws1, f"{report_label} BYOD SIM Fee Report", 5)
                subtitle_row(ws1, "Market-Level Summary  |  Sub Dealer & Preferred Only", 5)
                for ci, h in enumerate(['Nextar Market ID','Activations','SIM Fee','Returns','Net SIM Fee'], 1):
                    hdr(ws1.cell(3, ci), h, wrap=True)
                ws1.row_dimensions[3].height = 28
                row = 4
                for _, r in market_summary.iterrows():
                    bg = GRAY if row % 2 == 0 else WHITE
                    dcell(ws1.cell(row,1), r['Market'], bg=bg)
                    dcell(ws1.cell(row,2), int(r['Activations']), bg=bg, center=True)
                    dcell(ws1.cell(row,3), int(r['BYOD_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws1.cell(row,4), int(r['Returns']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws1.cell(row,5), int(r['Net_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    border_row(ws1, row, 1, 5); row += 1
                hdr(ws1.cell(row,1), 'TOTAL', bg=DARK_BLUE, size=10)
                dcell(ws1.cell(row,2), total_acts, bold=True, bg=GOLD, center=True)
                dcell(ws1.cell(row,3), total_fee, bold=True, bg=GOLD, center=True, num_fmt='$#,##0')
                dcell(ws1.cell(row,4), total_ret, bold=True, bg=GOLD, center=True, num_fmt='$#,##0')
                dcell(ws1.cell(row,5), total_net, bold=True, bg=GOLD, center=True, num_fmt='$#,##0')
                border_row(ws1, row, 1, 5)
                ws1.row_dimensions[row].height = 18
                for col, w in zip('ABCDE', [18,14,14,14,14]): ws1.column_dimensions[col].width = w

                # ── Sheet 2: Dealer ──
                ws2 = wb.create_sheet("Dealer"); ws2.sheet_view.showGridLines = False
                title_row(ws2, f"{report_label} BYOD SIM Fee Report", 9)
                subtitle_row(ws2, "Dealer-Level Summary — Sub Dealer & Preferred Only", 9)
                for ci, h in enumerate(['Market','Account Type','Dealer Name','JKME?','Doors','Activations','SIM Fee','Returns','Net SIM Fee'], 1):
                    hdr(ws2.cell(3, ci), h, wrap=True)
                ws2.row_dimensions[3].height = 28
                for ri, (_, dr) in enumerate(dealer_fees.iterrows(), 4):
                    bg = VERY_LIGHT if ri % 2 == 0 else WHITE
                    dcell(ws2.cell(ri,1), dr['Market'], bg=bg)
                    dcell(ws2.cell(ri,2), dr['Account Type'], bg=bg)
                    dcell(ws2.cell(ri,3), dr['Location Name'], bg=bg)
                    dcell(ws2.cell(ri,4), 'Yes' if dr['Is_JKME'] else 'No', bg=bg, center=True)
                    dcell(ws2.cell(ri,5), int(dr['Doors']), bg=bg, center=True)
                    dcell(ws2.cell(ri,6), int(dr['Activations']), bg=bg, center=True)
                    dcell(ws2.cell(ri,7), int(dr['Total_SIM_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws2.cell(ri,8), int(dr['Returns']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws2.cell(ri,9), int(dr['Net_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    border_row(ws2, ri, 1, 9)
                ws2.auto_filter.ref = f"A3:{get_column_letter(9)}{len(dealer_fees)+3}"
                ws2.freeze_panes = 'A4'
                for col, w in zip('ABCDEFGHI', [14,14,30,9,8,13,12,11,13]): ws2.column_dimensions[col].width = w

                # ── Sheet 3: Door ──
                ws3 = wb.create_sheet("Door"); ws3.sheet_view.showGridLines = False
                title_row(ws3, f"{report_label} BYOD SIM Fee Report", 13)
                subtitle_row(ws3, "Door-Level Summary — Sub Dealer & Preferred Only", 13)
                for ci, h in enumerate(['Door ID','Nextar Market ID','Account Type','Dealer Name','Street Address','City','State','ZIP','RT2 ID','Activations','SIM Fee','Returns','Net SIM Fee'], 1):
                    hdr(ws3.cell(3, ci), h, wrap=True)
                ws3.row_dimensions[3].height = 28
                for ri, (_, dr) in enumerate(door_fees.iterrows(), 4):
                    bg = VERY_LIGHT if ri % 2 == 0 else WHITE
                    dcell(ws3.cell(ri,1), dr['Door Code'], bg=bg)
                    dcell(ws3.cell(ri,2), dr['Market'], bg=bg)
                    dcell(ws3.cell(ri,3), dr['Account Type'], bg=bg)
                    dcell(ws3.cell(ri,4), dr['Location Name'], bg=bg)
                    dcell(ws3.cell(ri,5), dr['Street Address'], bg=bg)
                    dcell(ws3.cell(ri,6), dr['City'], bg=bg)
                    dcell(ws3.cell(ri,7), dr['State'], bg=bg, center=True)
                    dcell(ws3.cell(ri,8), dr['ZIP'], bg=bg, center=True)
                    dcell(ws3.cell(ri,9), dr['RT2 ID'], bg=bg)
                    dcell(ws3.cell(ri,10), int(dr['Activations']), bg=bg, center=True)
                    dcell(ws3.cell(ri,11), int(dr['BYOD_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws3.cell(ri,12), int(dr['Returns']), bg=bg, center=True, num_fmt='$#,##0')
                    dcell(ws3.cell(ri,13), int(dr['Total_Fee']), bg=bg, center=True, num_fmt='$#,##0')
                    border_row(ws3, ri, 1, 13)
                ws3.auto_filter.ref = f"A3:{get_column_letter(13)}{len(door_fees)+3}"
                ws3.freeze_panes = 'A4'
                for col, w in zip('ABCDEFGHIJKLM', [12,14,13,28,28,16,7,8,12,12,11,10,12]):
                    ws3.column_dimensions[col].width = w

                # ── Sheet 4: BYOD Detail ──
                ws4 = wb.create_sheet("BYOD"); ws4.sheet_view.showGridLines = False
                merged_sorted = merged.sort_values(['Market','Location Name','Door Code'])
                source_cols = ['AR Code','AR Name','Door Code','Door Name','Market','Account Type','Location Name',
                               'Door Address','Original Door Code','Original Door Name','EDGE Login','Employee ID',
                               'Employee Name','Account Number','Subscriber ID','MDN','IMEI','SIM','SIM History',
                               'IMEI History / MIM ACTV Code','T-Mo ESN History Date / MIM Order Date','RMA',
                               'Carrier Name','SKU','Handset Model','IMEI Trade-In / MIM Order ID',
                               'MIM/Traded-in Chargeback Reason/Comments','Transaction Date','Program Name',
                               'Transaction Type','Qualification Day','Qualification Status','Original SOC',
                               'Comp SOC','Comp MRC','Transaction Amount','Business Rule','Posted Date',
                               'ACH Cash or Credit','Transaction Amount2','Commission MRC','BYOD Fee']
                target_names = ['AR Code','AR Name','Door Code','Door Name','Nextar Market','Door Type','Dealer Name',
                                'Door Address','Original Door Code','Original Door Name','EDGE Login','Employee ID',
                                'Employee Name','Account Number','Subscriber ID','MDN','ESN','SIM','SIM History',
                                'IMEI History','TMO Network History Date','RMA','Carrier Name','SKU','Handset Model',
                                'IMEI Trade In','Traded-In Chargeback Reason / Comments','Transaction Date String',
                                'Program Name','Transaction Type','Qualification Day','Qualification Status',
                                'Original SOC','Comp SOC','Comp MRC','Transaction Amount','Business Rule',
                                'Posted Date','ACH Cash or Credit','Transaction Amount2','Commission MRC','BYOD Fee']
                numeric_cols = {'Comp MRC','Transaction Amount','Transaction Amount2','Commission MRC','BYOD Fee'}
                currency_cols = {'Transaction Amount2','Commission MRC','BYOD Fee'}

                detail_df = merged_sorted[[c for c in source_cols if c in merged_sorted.columns]].copy()
                detail_df.columns = target_names[:len(detail_df.columns)]
                ncols = len(detail_df.columns)
                last_col = get_column_letter(ncols)

                title_row(ws4, f"{report_label} BYOD — Transaction Detail  |  Sub Dealer & Preferred Only", ncols)
                for ci, col in enumerate(detail_df.columns, 1):
                    cell = ws4.cell(2, ci); cell.value = col
                    cell.font = Font(bold=True, size=9, color=WHITE, name='Arial')
                    cell.fill = PatternFill('solid', start_color=MED_BLUE)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws4.row_dimensions[2].height = 28
                for ri, (_, row_data) in enumerate(detail_df.iterrows(), 3):
                    bg = VERY_LIGHT if ri % 2 == 0 else WHITE
                    for ci, (col_name, val) in enumerate(zip(detail_df.columns, row_data), 1):
                        cell = ws4.cell(ri, ci)
                        is_byod = (ci == ncols)
                        is_num = col_name in numeric_cols
                        is_curr = col_name in currency_cols
                        if is_num and pd.notna(val):
                            try: cell.value = float(val)
                            except: cell.value = val if pd.notna(val) else ''
                        else:
                            cell.value = val if pd.notna(val) else ''
                        cell.font = Font(bold=is_byod, size=9, name='Arial', color='1F3864' if is_byod else '000000')
                        cell.fill = PatternFill('solid', start_color=bg)
                        cell.alignment = Alignment(vertical='center', horizontal='center' if is_num else 'left')
                        if is_curr: cell.number_format = '$#,##0.00'
                        elif is_num: cell.number_format = '#,##0.00'
                ws4.auto_filter.ref = f"A2:{last_col}{len(detail_df)+2}"
                ws4.freeze_panes = 'A3'
                col_widths = [14,22,11,24,10,10,28,28,12,12,10,10,18,13,13,13,18,22,10,12,12,10,10,16,22,16,16,13,14,18,10,12,12,12,10,14,12,13,12,14,12,10]
                for ci, w in enumerate(col_widths[:ncols], 1):
                    ws4.column_dimensions[get_column_letter(ci)].width = w

                # ── Sheet 5: Fee Schedule ──
                ws5 = wb.create_sheet("Fee Schedule"); ws5.sheet_view.showGridLines = False
                ws5.merge_cells('A1:C1')
                t5 = ws5['A1']; t5.value = "SIM Fee Schedule Reference"
                t5.font = Font(bold=True, size=14, color=WHITE, name='Arial')
                t5.fill = PatternFill('solid', start_color=DARK_BLUE)
                t5.alignment = Alignment(horizontal='center', vertical='center')
                ws5.row_dimensions[1].height = 28
                for ci, h in enumerate(['Comp Range (Transaction Amount2)','Door Type','SIM Fee'], 1):
                    hdr(ws5.cell(3, ci), h)
                fee_tiers = [
                    (f'< ${t1_max:.2f}', 'Standard', '$0', WHITE),
                    (f'${t1_max:.2f} – ${t2_max:.2f}', 'Standard', '$5', GRAY),
                    (f'${t2_max:.2f} – ${t3_max:.2f}', 'Standard', '$10', WHITE),
                    (f'> ${t3_max:.2f}', 'Standard', '$15', GRAY),
                    ('Any amount', 'JKME Doors', '$5', LIGHT_GREEN),
                ]
                for i, (comp, dtype, fee, bg) in enumerate(fee_tiers, 4):
                    jkme = 'JKME' in dtype
                    ws5.cell(i,1).value = comp; ws5.cell(i,1).fill = PatternFill('solid', start_color=bg)
                    ws5.cell(i,1).alignment = Alignment(horizontal='center', vertical='center')
                    ws5.cell(i,1).font = Font(name='Arial', size=10)
                    ws5.cell(i,2).value = dtype; ws5.cell(i,2).fill = PatternFill('solid', start_color=bg)
                    ws5.cell(i,2).alignment = Alignment(horizontal='center', vertical='center')
                    ws5.cell(i,2).font = Font(bold=jkme, name='Arial', size=10, color='375623' if jkme else '000000')
                    ws5.cell(i,3).value = fee; ws5.cell(i,3).fill = PatternFill('solid', start_color=bg)
                    ws5.cell(i,3).alignment = Alignment(horizontal='center', vertical='center')
                    ws5.cell(i,3).font = Font(bold=True, name='Arial', size=10)
                    border_row(ws5, i, 1, 3)
                ws5.merge_cells('A10:C10')
                note = ws5['A10']; note.value = "Return Credits"
                note.font = Font(bold=True, size=10, color=WHITE, name='Arial')
                note.fill = PatternFill('solid', start_color=DARK_BLUE)
                note.alignment = Alignment(horizontal='center', vertical='center')
                ws5.row_dimensions[10].height = 18
                hdr(ws5.cell(11,1), 'Device Type'); hdr(ws5.cell(11,2), 'Return Credit')
                ws5.cell(11,3).fill = PatternFill('solid', start_color=MED_BLUE)
                for row, dtype, credit, bg in [(12,'Phone',f'${phone_credit}',WHITE),(13,'BTS',f'${bts_credit}',GRAY)]:
                    ws5.cell(row,1).value = dtype; ws5.cell(row,1).fill = PatternFill('solid', start_color=bg)
                    ws5.cell(row,1).alignment = Alignment(horizontal='center', vertical='center')
                    ws5.cell(row,1).font = Font(name='Arial', size=10)
                    ws5.cell(row,2).value = credit; ws5.cell(row,2).font = Font(bold=True, name='Arial', size=10)
                    ws5.cell(row,2).fill = PatternFill('solid', start_color=bg)
                    ws5.cell(row,2).alignment = Alignment(horizontal='center', vertical='center')
                    border_row(ws5, row, 1, 2)
                for col, w in zip('ABC', [32,18,12]): ws5.column_dimensions[col].width = w

                # ── Sheet 6: COO (only if exceptions exist) ──
                if coo_log:
                    ws6 = wb.create_sheet("COO Summary"); ws6.sheet_view.showGridLines = False
                    coo_headers = ['Door Code','Old Dealer','Old Account Type','New Dealer','New Account Type',
                                   'Effective Date','Txns Before (Old Owner)','Txns After (New Owner)',
                                   'Old Owner Charged?','New Owner Charged?']
                    title_row(ws6, f"{report_label} — Change of Ownership Summary", len(coo_headers))
                    subtitle_row(ws6, "Transactions split by ownership change date", len(coo_headers))
                    for ci, h in enumerate(coo_headers, 1):
                        hdr(ws6.cell(3, ci), h, wrap=True)
                    ws6.row_dimensions[3].height = 36
                    for ri, entry in enumerate(coo_log, 4):
                        bg = VERY_LIGHT if ri % 2 == 0 else WHITE
                        old_charged = 'Yes' if entry['Old Type'].lower() in ['sub','preferred'] else 'No — Direct'
                        new_charged = 'Yes' if entry['New Type'].lower() in ['sub','preferred'] else 'No — Direct'
                        vals = [entry['Door Code'], entry['Old Dealer'], entry['Old Type'],
                                entry['New Dealer'], entry['New Type'], entry['Effective Date'],
                                entry['Txns Before (Old Owner)'], entry['Txns After (New Owner)'],
                                old_charged, new_charged]
                        for ci, v in enumerate(vals, 1):
                            dcell(ws6.cell(ri, ci), v, bg=bg,
                                  center=(ci in [7,8]),
                                  fg=('375623' if 'Yes' in str(v) else ('721C24' if 'No' in str(v) else '000000')))
                        border_row(ws6, ri, 1, len(coo_headers))
                    ws6.auto_filter.ref = f"A3:{get_column_letter(len(coo_headers))}{len(coo_log)+3}"
                    ws6.freeze_panes = 'A4'
                    for col, w in zip('ABCDEFGHIJ', [12,28,16,28,16,14,18,18,16,16]):
                        ws6.column_dimensions[col].width = w

                # ── Save to buffer ──
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                fname = f"{report_month}_{report_year}_BYOD_SIM_Fee_Report.xlsx"
                st.success(f"✅ Report generated successfully!")
                st.download_button(
                    label=f"⬇️ Download {fname}",
                    data=buffer,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"❌ Error generating report: {str(e)}")
                st.exception(e)

    st.markdown('</div>', unsafe_allow_html=True)
